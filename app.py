import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, colors, Font, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import matplotlib.colors as mcolors

st.set_page_config(
    page_title="Option Chain Formatter",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
    .main { max-width: 1000px; padding: 2rem; }
    .stButton>button { width: 100%; padding: 0.5rem; border-radius: 5px; }
    .stDownloadButton>button { background-color: #4CAF50; color: white; }
    </style>
""", unsafe_allow_html=True)


def process_option_chain(df):
    try:
        raw = df.copy()
        headers = raw.iloc[1].tolist()
        headers = [str(h).strip() if pd.notna(h) and str(h).strip() != '' else f"Column_{i}" for i, h in
                   enumerate(headers)]
        data = raw.iloc[2:].copy()
        data.columns = headers
        data = data.reset_index(drop=True)

        def col_positions(name: str):
            return [i for i, v in enumerate(headers) if str(v).strip() == name]

        def series_by_pos(name: str, occurrence: int = 0) -> pd.Series:
            pos_list = col_positions(name)
            if len(pos_list) <= occurrence:
                raise KeyError(f"Column '{name}' occurrence {occurrence} not found.")
            return data.iloc[:, pos_list[occurrence]]

        time_s = series_by_pos("Time", 0)
        strike_s = pd.to_numeric(series_by_pos("Strike Price", 0), errors="coerce")
        ce_oi_chg_s = pd.to_numeric(series_by_pos("OI Chg", 0), errors="coerce")
        ce_oi_s = pd.to_numeric(series_by_pos("OI", 0), errors="coerce")
        ce_vwap_s = pd.to_numeric(series_by_pos("VWAP", 0), errors="coerce")
        pe_vwap_s = pd.to_numeric(series_by_pos("VWAP", 1), errors="coerce")
        pe_oi_s = pd.to_numeric(series_by_pos("OI", 1), errors="coerce")
        pe_oi_chg_s = pd.to_numeric(series_by_pos("OI Chg", 1), errors="coerce")

        out = pd.DataFrame()
        out["Time"] = time_s
        out["CE OI CHANGE"] = ce_oi_chg_s
        out["CE OI"] = ce_oi_s
        out["CE MONEY"] = ((out["CE OI CHANGE"] * ce_vwap_s) / 10_000_000).round(0)
        out["CE BEP"] = (strike_s + ce_vwap_s).round(0)
        out["CE VWAP"] = ce_vwap_s
        out["Strike Price"] = strike_s
        out["PE VWAP"] = pe_vwap_s
        out["PE BEP"] = (strike_s - pe_vwap_s).round(0)
        out["PE MONEY"] = ((pe_oi_chg_s * pe_vwap_s) / 10_000_000).round(0)
        out["PE OI"] = pe_oi_s
        out["PE OI CHANGE"] = pe_oi_chg_s

        final_order = [
            "Time",
            "CE OI CHANGE", "CE OI", "CE MONEY", "CE BEP", "CE VWAP",
            "Strike Price",
            "PE VWAP", "PE BEP", "PE MONEY", "PE OI", "PE OI CHANGE"
        ]

        return out[final_order]

    except Exception as e:
        import traceback
        st.error(f"Error processing file: {str(e)}")
        st.error(f"Error details: {traceback.format_exc()}")
        return None


def apply_formatting(worksheet, df):
    worksheet.freeze_panes = "A2"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    oi_change_cols = {'B': 'CE OI CHANGE', 'K': 'PE OI CHANGE'}
    for col_letter, col_name in oi_change_cols.items():
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None and cell.value < 0:
                    cell.font = Font(color='FF0000')

    for col in ['B', 'L']:
        color_scale_rule = ColorScaleRule(
            start_type='min',
            start_color='F8696B', mid_type='num', mid_value=0,
            mid_color='FFEB84', end_type='max', end_color='63BE7B'
        )
        worksheet.conditional_formatting.add(f'{col}2:{col}{len(df) + 1}', color_scale_rule)

    for col in ['D', 'J']:
        color_scale_rule = ColorScaleRule(
            start_type='min',
            start_color='9DC3E6', mid_type='percentile', mid_value=50,
            mid_color='FFFFFF', end_type='max', end_color='1F4E79'
        )
        worksheet.conditional_formatting.add(f'{col}2:{col}{len(df) + 1}', color_scale_rule)
    # Color scaling for CE OI and PE OI columns (C and K)
    for col in ['C', 'K']:
        color_scale_rule = ColorScaleRule(
            start_type='min',
            start_color='FFF2CC',  # Light Beige for min
            mid_type='percentile',
            mid_value=50,
            mid_color='F4B183',  # Orange for median
            end_type='max',
            end_color='8B4513'  # Dark Brown for max
        )
        worksheet.conditional_formatting.add(f'{col}2:{col}{len(df) + 1}', color_scale_rule)
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 25)
    # Define color fills
    ce_bep_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')  # Light Orange
    pe_bep_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light Green

    # Top 3 indices of each metric
    top_ce_oi_chg_idx = set(df['CE OI CHANGE'].nlargest(4).index.tolist())
    top_ce_money_idx = set(df['CE MONEY'].nlargest(4).index.tolist())
    top_pe_oi_chg_idx = set(df['PE OI CHANGE'].nlargest(4).index.tolist())
    top_pe_money_idx = set(df['PE MONEY'].nlargest(4).index.tolist())

    # Intersections for coloring
    top_ce_indices = top_ce_oi_chg_idx.intersection(top_ce_money_idx)
    top_pe_indices = top_pe_oi_chg_idx.intersection(top_pe_money_idx)

    # Color CE BEP cells
    for idx in top_ce_indices:
        worksheet.cell(row=idx + 2, column=5).fill = ce_bep_fill  # CE BEP is column D

    # Color PE BEP cells
    for idx in top_pe_indices:
        worksheet.cell(row=idx + 2, column=9).fill = pe_bep_fill  # PE BEP is column I

    return worksheet


def style_dataframe_for_preview(df):
    styled_df = df.style

    # Exact color scale mappings used in Excel export
    cmap_oi_change = mcolors.LinearSegmentedColormap.from_list(
        "oi_change", ['#F8696B', '#FFEB84', '#63BE7B']
    )
    cmap_oi = mcolors.LinearSegmentedColormap.from_list(
        "oi", ['#FFF2CC', '#F4B183', '#786C3B']
    )
    cmap_money = mcolors.LinearSegmentedColormap.from_list(
        "money", ['#9DC3E6', '#FFFFFF', '#1F4E79']
    )

    columns_to_style = {
        'CE OI CHANGE': cmap_oi_change,
        'PE OI CHANGE': cmap_oi_change,
        'CE OI': cmap_oi,
        'PE OI': cmap_oi,
        'CE MONEY': cmap_money,
        'PE MONEY': cmap_money
    }

    for col, cmap in columns_to_style.items():
        if col in df.columns:
            styled_df = styled_df.background_gradient(
                cmap=cmap,
                subset=[col]
            )

    # Top 3 indices of each metric
    top_ce_oi_chg_idx = set(df['CE OI CHANGE'].nlargest(4).index.tolist())
    top_ce_money_idx = set(df['CE MONEY'].nlargest(4).index.tolist())
    top_pe_oi_chg_idx = set(df['PE OI CHANGE'].nlargest(4).index.tolist())
    top_pe_money_idx = set(df['PE MONEY'].nlargest(4).index.tolist())

    top_ce_indices = top_ce_oi_chg_idx.intersection(top_ce_money_idx)
    top_pe_indices = top_pe_oi_chg_idx.intersection(top_pe_money_idx)

    # Apply custom highlight
    def highlight_bep(row):
        styles = ['' for _ in row]
        if row.name in top_ce_indices:
            styles[df.columns.get_loc('CE BEP')] = 'background-color: #FFD966'  # Light orange
        if row.name in top_pe_indices:
            styles[df.columns.get_loc('PE BEP')] = 'background-color: #C6EFCE'  # Light green
        return styles

    styled_df = styled_df.apply(highlight_bep, axis=1)

    return styled_df


def main():
    st.title("📊 Option Chain Formatter")
    st.markdown("---")
    st.write("Upload your option chain Excel file to process and download the formatted version.")

    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

    if uploaded_file is not None:
        try:
            raw_df = pd.read_excel(uploaded_file, header=None, dtype=str)
            with st.spinner('Processing your file...'):
                processed_df = process_option_chain(raw_df)
                if processed_df is not None:
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        processed_df.to_excel(writer, index=False, sheet_name='OptionChain')
                        worksheet = writer.sheets['OptionChain']
                        apply_formatting(worksheet, processed_df)

                    st.success("✅ Processing complete!")

                    import os
                    base_name = os.path.splitext(uploaded_file.name)[0]
                    output_filename = f"{base_name}_processed.xlsx"

                    st.download_button(
                        label="⬇️ Download Processed File",
                        data=output.getvalue(),
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    st.subheader("Preview of Processed Data")
                    pd.set_option('display.max_rows', None)
                    st.dataframe(style_dataframe_for_preview(processed_df))
                    pd.reset_option('display.max_rows')

        except Exception as e:
            st.error(f"❌ An error occurred: {str(e)}")

    with st.sidebar:
        st.title("ℹ️ Instructions")
        st.markdown("""
        1. Upload your option chain Excel file (iCharts format ONLY supported).
        2. Wait for processing.
        3. Download the formatted file or Preview it.
        """)
        st.markdown("---")
        st.markdown("### About")
        st.markdown("This tool formats option chain data to calculate BEPs and OI Money.")
        st.markdown("© 2025 Manikanth Mulukutla. All rights reserved.")


if __name__ == "__main__":
    main()
