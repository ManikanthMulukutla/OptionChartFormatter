import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl.styles import PatternFill, colors
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font, Border, Side, Alignment

APP_TITLE = "Option Chain Formatter"

def process_option_chain(input_path: str, output_path: str):
    """
    Reads the input Excel (with the 2-row header pattern you shared),
    selects & renames columns, computes CE/PE MONEY & BEP, and saves output.
    """
    # 1) Read raw with no header so we can control header rows
    raw = pd.read_excel(input_path, header=None)

    # 2) Row 1 (0-based) contains the real headers; data starts at row 2
    headers = raw.iloc[1].tolist()
    data = raw.iloc[2:].copy()
    data.columns = headers

    # Helpers to handle duplicate column names (e.g., VWAP, OI, OI Chg appear twice)
    def col_positions(name: str):
        return [i for i, v in enumerate(headers) if isinstance(v, str) and v.strip() == name]

    def series_by_pos(name: str, occurrence: int = 0) -> pd.Series:
        pos_list = col_positions(name)
        if len(pos_list) <= occurrence:
            raise KeyError(f"Column '{name}' occurrence {occurrence} not found.\n"
                           f"Available positions for '{name}': {pos_list}")
        return data.iloc[:, pos_list[occurrence]]

    # 3) Map all needed columns (CE = first occurrence, PE = second)
    time_s       = series_by_pos("Time", 0)                # left time
    strike_s     = pd.to_numeric(series_by_pos("Strike Price", 0), errors="coerce")

    ce_oi_chg_s  = pd.to_numeric(series_by_pos("OI Chg", 0), errors="coerce")
    ce_oi_s      = pd.to_numeric(series_by_pos("OI", 0), errors="coerce")
    ce_vwap_s    = pd.to_numeric(series_by_pos("VWAP", 0), errors="coerce")

    pe_vwap_s    = pd.to_numeric(series_by_pos("VWAP", 1), errors="coerce")
    pe_oi_s      = pd.to_numeric(series_by_pos("OI", 1), errors="coerce")
    pe_oi_chg_s  = pd.to_numeric(series_by_pos("OI Chg", 1), errors="coerce")

    # 4) Build output with your exact order + computed fields
    out = pd.DataFrame()
    out["Time"]           = time_s
    out["CE OI CHANGE"]   = ce_oi_chg_s
    out["CE OI"]          = ce_oi_s
    out["CE MONEY"]       = ((out["CE OI CHANGE"] * ce_vwap_s) / 10_000_000).round(0)
    out["CE BEP"]         = (strike_s + ce_vwap_s).round(0)
    out["CE VWAP"]        = ce_vwap_s
    out["Strike Price"]   = strike_s
    out["PE VWAP"]        = pe_vwap_s
    out["PE BEP"]         = (strike_s - pe_vwap_s).round(0)
    out["PE MONEY"]       = ((pe_oi_chg_s * pe_vwap_s) / 10_000_000).round(0)
    out["PE OI"]          = pe_oi_s
    out["PE OI CHANGE"]   = pe_oi_chg_s

    final_order = [
        "Time",
        "CE OI CHANGE",
        "CE OI",
        "CE MONEY",
        "CE BEP",
        "CE VWAP",
        "Strike Price",
        "PE VWAP",
        "PE BEP",
        "PE MONEY",
        "PE OI",
        "PE OI CHANGE",
    ]
    out = out[final_order]

    # 5) Apply color scaling and save
    def save_with_color_scaling(df, output_path):
        # Save to Excel first
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='OptionChain')
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['OptionChain']
        
        # Define styles
        font_style = Font(size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply styles to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font_style
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths after setting font
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Add some extra width for the larger font
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 20)  # Increased max width
        
        # Make header row bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=14)
        
        # Define color scale rules
        # For OI Change columns (red for negative, green for positive)
        oi_change_cols = ['B', 'L']  # CE OI CHANGE and PE OI CHANGE
        for col in oi_change_cols:
            # Green for positive values
            color_scale_rule = ColorScaleRule(
                start_type='min', start_color='9DC3E6',  # Light Blue for min
                mid_type='num', mid_value=0, mid_color='FFFFFF',  # White for zero
                end_type='max', end_color='1F4E79'  # Green for max
            )
            worksheet.conditional_formatting.add(f'{col}2:{col}1000', color_scale_rule)
        
        # For Money columns (gradient from red to green)
        money_cols = ['D', 'J']  # CE MONEY and PE MONEY
        for col in money_cols:
            # Color scale from red (negative) to green (positive)
            color_scale_rule = ColorScaleRule(
                start_type='min', start_color='F8696B',  # Red for min
                mid_type='num', mid_value=0, mid_color='FFEB84',  # White for zero
                end_type='max', end_color='63BE7B'  # Green for max
            )
            worksheet.conditional_formatting.add(f'{col}2:{col}1000', color_scale_rule)
        
        # Save the workbook with all formatting applied
        workbook.save(output_path)
    
    # Save with color scaling
    save_with_color_scaling(out, output_path)

def run_app():
    root = tk.Tk()
    root.title(APP_TITLE)
    # Increased window size to 600x300
    root.geometry("600x300")

    def choose_and_process():
        in_path = filedialog.askopenfilename(
            title="Select Input Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not in_path:
            return

        default_name = f"OptionChain_Processed.xlsx"
        out_path = filedialog.asksaveasfilename(
            title="Save Processed Excel As",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not out_path:
            return

        try:
            process_option_chain(in_path, out_path)
            messagebox.showinfo(APP_TITLE, f"Success!\nSaved to:\n{out_path}")
        except KeyError as e:
            messagebox.showerror(APP_TITLE, f"Missing expected columns:\n{e}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Error:\n{e}")

    # Create a frame for centering content
    content_frame = tk.Frame(root)
    content_frame.pack(expand=True, fill='both', padx=20, pady=20)
    
    # Center the heading and button
    tk.Label(content_frame, text="Option Chain → Formatter", 
             font=("Segoe UI", 16, "bold")).pack(pady=(20, 10), anchor='center')
    
    tk.Label(content_frame, 
             text="Pick the input Excel, I'll compute MONEY & BEP and save the output.",
             font=("Segoe UI", 10)).pack(pady=(0, 20), anchor='center')
    
    button_frame = tk.Frame(content_frame)
    button_frame.pack(expand=True, pady=10)
    
    tk.Button(button_frame, text="Select & Process Excel",
              font=("Segoe UI", 12),
              command=choose_and_process).pack(pady=10)
    
    # Add copyright notice
    tk.Label(root, text=" 2025 Manikanth Mulukutla. All rights reserved.",
             font=("Segoe UI", 8), fg="gray").pack(side="bottom", pady=5)

    root.mainloop()

if __name__ == "__main__":
    run_app()
